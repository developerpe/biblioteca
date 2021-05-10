from datetime import datetime

from django.http import HttpResponse
from django.views.generic import TemplateView

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,Border,Font,Side
)
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from apps.automatic_crud.generics import BaseCrudMixin
from apps.automatic_crud.utils import (
    get_model,get_model_fields_names,get_queryset
)

def _excel_report_title(__model_name: str):
    """
    Build report title with today date
    """


    date = datetime.now()
    title = "REPORTE DE {0} EN FORMATO EXCEL REALIZADO EN LA FECHA: {1}".format(
                                                                            __model_name.upper(),
                                                                            "%s/%s/%s" % (
                                                                                date.day,date.month,
                                                                                date.year
                                                                            )
                                                                        )
    return title

def _validate_id(__field: str):
    if str(__field).lower() != 'id':
        return True
    return False

class ExcelReportFormat:
    """
    This class generates a report in excel for any model you want, 
    the parameters to be used are defined in the constructor, 
    and there are also some methods that build the report block by block.

    Parameters:
        _app_name                   name of the application where is the model to be used.
        _model_name                 name of the model to be used.

    Variables:
        _app_name                   name of the application where is the model to be used.
        _model_name                 name of the model to be used.
        __model                     model to be used
        __model_fields_names        fields list of model.
        __queryset                  queryset of model, contains all registers of model.
        __report_title              report title.
        __workbook                  Workbook instance, Excel workbook.
        __sheetwork                 Excel Sheetwork, by default first sheet.

    """
    

    def __init__(self,__app_name:str,__model_name:str, *args, **kwargs):
        self.__app_name = __app_name
        self.__model_name = __model_name
        self.__model = get_model(self.__app_name,self.__model_name)
        self.__model_fields_names = get_model_fields_names(self.__model)
        self.__queryset = get_queryset(self.__model)
        self.__report_title = _excel_report_title(self.__model_name)
        self.__workbook = Workbook()
        self.__sheetwork = self.__workbook.active

    def get_model(self):
        return self.__model

    def __excel_report_header(self,row_dimension = 15, col_dimension = 25):
        """
        Build excel report header, print report title and add default styles

        """

        self.__sheetwork['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
        self.__sheetwork['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        self.__sheetwork['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
        self.__sheetwork['B1'] = self.__report_title
        
        if len(self.__model_fields_names) < 12:
            __header_letter = 'L'
        else:
            __header_letter = '{0}'.format(get_column_letter(len(self.__model_fields_names)).upper())
        
        self.__sheetwork.merge_cells('B1:{0}1'.format(__header_letter))
        self.__sheetwork.row_dimensions[3].height = row_dimension

        __count = 1
        for __field in self.__model_fields_names:
            __letter = get_column_letter(__count).upper()
            self.__sheetwork['{0}3'.format(__letter)].alignment = Alignment(horizontal = "center", vertical = "center")
            self.__sheetwork['{0}3'.format(__letter)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            self.__sheetwork['{0}3'.format(__letter)].font = Font(name = 'Calibri', size = 9, bold = True)
            self.__sheetwork['{0}3'.format(__letter)] = '{0}'.format(__field.upper())
            self.__sheetwork.column_dimensions['{0}'.format(__letter)].height = col_dimension
            __count += 1
    
    def __print_values(self):
        """
        Print values of queryset, skip id value and exclude fields of model
        """

        row_count = 4
        col_count = 1
        general_count = len(self.__model_fields_names)
        data = [value for value in self.__queryset]

        for value in data:
            for key,subvalue in value.items():
                if _validate_id(key):
                    if key not in self.__model.exclude_fields:
                        if general_count > 0:
                            self.__sheetwork.cell(row = row_count, column = col_count).alignment = Alignment(horizontal = "center")
                            self.__sheetwork.cell(row = row_count, column = col_count).border = Border(left = Side(border_style = "thin"),
                                                                        right = Side(border_style = "thin"),top = Side(border_style = "thin"), 
                                                                        bottom = Side(border_style = "thin"))
                            if type(subvalue) is bool:
                                if subvalue is True:
                                    self.__sheetwork.cell(row = row_count, column = col_count).value = 'No eliminado'
                                else:
                                    self.__sheetwork.cell(row = row_count, column = col_count).value = 'Eliminado'
                            else:
                                self.__sheetwork.cell(row = row_count, column = col_count).value = str(subvalue)
                                if (self.__sheetwork.column_dimensions[get_column_letter(col_count).upper()].width < len(str(subvalue))):
                                    self.__sheetwork.column_dimensions[get_column_letter(col_count).upper()].width = len(str(subvalue))
                            
                            col_count += 1
                            general_count -= 1

                    if general_count == 0:
                        general_count = len(self.__model_fields_names)
            
            row_count += 1
            col_count = 1

    def get_excel_report(self):
        """
        Generate excel response using model name
        """

        report_name = "Reporte {0} en Excel .xlsx".format(self.__model_name)
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(report_name)
        response['Content-Disposition'] = content
        self.__workbook.save(response)
        return response

    def build_report(self):
        """
        Build report call 2 functions: __excel_report_header and __print_values
        """

        self.__excel_report_header()
        self.__print_values()

class GetExcelReport(BaseCrudMixin,TemplateView):
    """
    Return Instance Excel Report for a model.
    """

    def get(self,request,_app_name:str,_model_name:str,*args,**kwargs):
        __report = ExcelReportFormat(_app_name,_model_name)

        self.model = __report.get_model()

        # login required validation
        validation_login_required,response = self.validate_login_required()
        if validation_login_required:
            return response
        
        # permission required validation
        validation_permissions,response = self.validate_permissions()
        if validation_permissions:
            return response

        __report.build_report()
        return __report.get_excel_report()